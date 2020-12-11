import openpyxl as xl
wb = xl.load_workbook('transaction.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
# print(sheet.max_row)

for row in range(1, sheet.max_row + 1):
    print(row)